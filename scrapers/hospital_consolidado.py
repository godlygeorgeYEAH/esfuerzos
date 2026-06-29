"""
scrapers/hospital_consolidado.py -- Registro Maestro de Pacientes (SISMO 2026).

Source: a community-maintained, continuously-updated Excel workbook consolidating
patients and survivors located in hospitals and shelters after the June 2026
Venezuela earthquake. One tab per hospital / refugio, columns:
  N° | APELLIDOS Y NOMBRES | EDAD | CÉDULA / ID | TELÉFONO | DIRECCIÓN | OBSERVACIONES

Why this source matters: it is the highest-signal "found" feed we have. Every row
is a person physically located (kind='found'), tagged with the hospital/shelter
(last_seen_location), and -- crucially -- many carry a CÉDULA. The cédula is copied
into distinguishing_marks as "CI: <digits>" so run_cedula_exact_match() can produce
exact-ID matches (the strongest possible confirmation) against family "missing" reports.

The workbook is a Dropbox shared file; the same shared URL serves the latest version
when the maintainers overwrite it, so a periodic re-download picks up updates. The
master "🔍 BUSCAR PACIENTES" tab is skipped (it aggregates the per-hospital tabs and
lacks a per-row hospital), avoiding double ingestion.

Config:
  HOSPITAL_XLSX_URL -- override the workbook URL (defaults to the known Dropbox dl=1 link).

Dependency: openpyxl (added to requirements.txt). The image must be rebuilt
(docker compose up --build) the first time this scraper ships.
"""
from __future__ import annotations

import io
import logging
import os
import re
from typing import Any

import httpx

from .base import BaseVEScraper

logger = logging.getLogger(__name__)

_SOURCE_NAME = "hospital_consolidado"

_DEFAULT_URL = (
    "https://dl.dropboxusercontent.com/scl/fi/m4fbaw4metvkuay91fi0j/"
    "26JUN26-15.12-Pacientes-Consolidados-Hospitales-Venezuela.xlsx"
    "?rlkey=0bjem2yymn9q88qumzr33fisz&dl=1"
)
_XLSX_URL = os.environ.get("HOSPITAL_XLSX_URL", _DEFAULT_URL)

# Tabs that are not per-hospital patient lists.
_SKIP_TABS = {"🔍 buscar pacientes", "buscar pacientes"}

# Header synonyms -> canonical field. Matched on a deaccented, lowercased cell.
_COL_NAME = ("apellidos y nombres", "nombres y apellidos", "nombre", "paciente")
_COL_AGE = ("edad",)
_COL_ID = ("cedula", "cedula / id", "cedula/id", "ci", "id")
_COL_PHONE = ("telefono", "teléfono", "contacto")
_COL_ADDR = ("direccion", "dirección")
_COL_OBS = ("observaciones", "observacion", "notas", "estado")

_MAX_MARKS = 480


def _deaccent(s: str) -> str:
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", (s or "").lower())
                   if unicodedata.category(c) != "Mn").strip()


def _match_col(cell: str, names: tuple[str, ...]) -> bool:
    c = _deaccent(cell)
    return any(c == n or c.startswith(n) for n in names)


def _digits(s: Any) -> str:
    return re.sub(r"\D", "", str(s or ""))


def _age_int(v: Any) -> int | None:
    m = re.search(r"\d{1,3}", str(v or ""))
    if not m:
        return None
    n = int(m.group(0))
    return n if 0 < n < 120 else None


def _header_index(rows: list[tuple], max_scan: int = 6) -> tuple[int, dict[str, int]]:
    """Find the header row (the one containing the name column) and map fields to
    column indices. Returns (header_row_idx, {field: col_idx}) or (-1, {})."""
    for i, row in enumerate(rows[:max_scan]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(_match_col(c, _COL_NAME) for c in cells):
            continue
        idx: dict[str, int] = {}
        for j, c in enumerate(cells):
            if "name" not in idx and _match_col(c, _COL_NAME):
                idx["name"] = j
            elif "age" not in idx and _match_col(c, _COL_AGE):
                idx["age"] = j
            elif "id" not in idx and _match_col(c, _COL_ID):
                idx["id"] = j
            elif "phone" not in idx and _match_col(c, _COL_PHONE):
                idx["phone"] = j
            elif "addr" not in idx and _match_col(c, _COL_ADDR):
                idx["addr"] = j
            elif "obs" not in idx and _match_col(c, _COL_OBS):
                idx["obs"] = j
        if "name" in idx:
            return i, idx
    return -1, {}


def _tab_location(rows: list[tuple], tab_name: str) -> str:
    """The hospital/shelter name: the title row (row 0, first non-empty cell)
    falls back to the tab name."""
    for row in rows[:2]:
        for c in row:
            if c and str(c).strip():
                title = str(c).strip()
                if len(title) > 3 and "buscar" not in _deaccent(title):
                    return title
            break
    return tab_name


def parse_workbook(data: bytes) -> list[dict]:
    """Pure parse: workbook bytes -> list of report dicts (no network/DB).

    Each per-hospital/shelter tab becomes a set of kind='found' reports located at
    that hospital, with the cédula embedded in distinguishing_marks for exact match.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict] = []
    for tab in wb.sheetnames:
        if _deaccent(tab) in _SKIP_TABS:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        location = _tab_location(rows, tab)
        hdr_i, idx = _header_index(rows)
        if hdr_i < 0:
            continue
        for row in rows[hdr_i + 1:]:
            if not row:
                continue
            def cell(field: str) -> str:
                j = idx.get(field)
                if j is None or j >= len(row):
                    return ""
                v = row[j]
                return str(v).strip() if v is not None else ""
            name = cell("name")
            if not name or len(name) < 3 or _deaccent(name) in ("n", "nombre", "total"):
                continue
            cedula = _digits(cell("id"))
            phone = cell("phone")
            addr = cell("addr")
            obs = cell("obs")
            marks_parts: list[str] = []
            if cedula and 5 <= len(cedula) <= 10:
                marks_parts.append(f"CI: {cedula}")          # feeds cedula exact match
            if obs:
                marks_parts.append(obs)
            if addr:
                marks_parts.append(f"Dir: {addr}")
            marks = " | ".join(marks_parts) or None
            if marks and len(marks) > _MAX_MARKS:
                marks = marks[:_MAX_MARKS - 3] + "..."
            # Status: the source marks deceased patients (fallecido/a). Flag it so
            # matching/notification handle it with care and the reviewer sees it
            # before confirming. Family closure ("vivo o muerto") needs this, but a
            # death must never be delivered as "encontrado sano".
            obs_d = _deaccent(obs)
            if any(t in obs_d for t in ("fallec", "muert", "occiso", "difunt")):
                person_state = "deceased"
            elif any(t in obs_d for t in ("alta", "egres")):
                person_state = "discharged"
            else:
                person_state = "found"
            # Stable per-row key: prefer cédula, else tab + N° + name slug.
            key = cedula or f"{_deaccent(tab)}:{_digits(cell('name')) or _deaccent(name)[:24]}"
            out.append({
                "kind": "found",
                "full_name": name,
                "age": _age_int(cell("age")),
                "last_seen_location": location,
                "distinguishing_marks": marks,
                "clothing": None,
                "person_state": person_state,
                "source": _SOURCE_NAME,
                "source_url": f"hospital_consolidado:{key}",
                "raw_data": {"hospital": location, "cedula": cedula or None,
                             "telefono": phone or None, "observaciones": obs or None,
                             "person_state": person_state},
            })
    return out


class HospitalConsolidadoScraper(BaseVEScraper):
    """Periodic ingester for the consolidated hospital/shelter patient workbook."""

    source_name = _SOURCE_NAME

    async def _download(self) -> bytes:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as cl:
            r = await cl.get(_XLSX_URL, headers={"User-Agent": "Mozilla/5.0 (ReuneVE)"})
            r.raise_for_status()
            return r.content

    async def _upsert_batch(self, rows: list[dict]) -> int:
        """Batch upsert into reports (chunks), merge on (source, source_url)."""
        from .base import _SUPABASE_URL, _SUPABASE_KEY
        if not _SUPABASE_URL or not _SUPABASE_KEY:
            raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set")
        done = 0
        async with httpx.AsyncClient(timeout=30) as cl:
            for i in range(0, len(rows), 500):
                chunk = rows[i:i + 500]
                resp = await cl.post(
                    f"{_SUPABASE_URL}/rest/v1/reports",
                    headers=self._sb_headers(),
                    params={"on_conflict": "source,source_url"},
                    json=chunk,
                )
                if resp.status_code in (200, 201, 204):
                    done += len(chunk)
                else:
                    logger.warning("hospital_consolidado upsert %d: %s",
                                   resp.status_code, resp.text[:160])
        return done

    async def poll_recent(self) -> int:
        # The workbook has no per-row recency and changes slowly; the hourly
        # full_sweep covers it. Skip the 5-minute poll to avoid re-parsing 13k rows.
        return 0

    async def full_sweep(self) -> int:
        total = 0
        error: str | None = None
        try:
            data = await self._download()
            reports = parse_workbook(data)
            logger.info("hospital_consolidado: parsed %d rows from workbook", len(reports))
            total = await self._upsert_batch(reports)
        except Exception as exc:  # noqa: BLE001
            error = str(exc)
            logger.error("hospital_consolidado full_sweep error: %s", exc)
        await self.log_run(_SOURCE_NAME, "full_sweep", total, 0, error)
        return total
